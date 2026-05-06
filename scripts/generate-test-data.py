#!/usr/bin/env python3
"""
Generate realistic test data for the dashboard demo.

Outputs four files matching the Microsoft Forms export shapes:
- templates/test-data/visitor-test-data.tsv      (~40 rows, paste-friendly)
- templates/test-data/outreach-test-data.tsv     (~12 rows, paste-friendly)
- templates/test-data/Visitor Test Data.xlsx     (Power Query-friendly)
- templates/test-data/Outreach Test Data.xlsx    (Power Query-friendly)

Use the .xlsx versions if the live Forms responses files are giving you
trouble — temporarily repoint your Power Query at the test files, refresh,
demo, then point Power Query back at the real Forms responses.

Distribution mirrors what you'd plausibly see at Conestoga's welcome desks:
- Doon busiest, then Waterloo, Reuter, Cambridge
- Afternoons heaviest, evenings sparsest
- Wayfinding the dominant inquiry category
- Categories sum >= headcount (one person can have multiple inquiries)

Deterministic via random.seed(42) so re-runs produce identical data.

Usage (from repo root):
    python scripts/generate-test-data.py
"""

import csv
import random
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


random.seed(42)


CAMPUSES = ["Waterloo", "Doon", "Reuter", "Cambridge"]
CAMPUS_WEIGHTS = [0.25, 0.40, 0.20, 0.15]

# Two time blocks (Morning open–12, Afternoon 12–close). Match exact wording
# in the form so SUMIFS-with-wildcard heatmap formulas catch them.
TIME_BLOCKS = [
    "Morning (open – 12pm)",
    "Afternoon (12 – close)",
]
TIME_BLOCK_WEIGHTS = [0.40, 0.60]
TIME_BLOCK_HOURS = {
    "Morning (open – 12pm)": (8, 12),
    "Afternoon (12 – close)": (12, 19),
}

VOLUNTEERS = [
    ("Maya Patel", "maya.patel@conestogac.on.ca"),
    ("Jordan Lee", "jordan.lee@conestogac.on.ca"),
    ("Aaron Chen", "aaron.chen@conestogac.on.ca"),
    ("Priya Singh", "priya.singh@conestogac.on.ca"),
    ("Sam Rivera", "sam.rivera@conestogac.on.ca"),
    ("Lucas Tremblay", "lucas.tremblay@conestogac.on.ca"),
    ("Fatima Khan", "fatima.khan@conestogac.on.ca"),
]

CATEGORIES = [
    "Wayfinding",
    "OneCard",
    "IT Support",
    "Bus Pass / Transportation",
    "Parking",
    "Timetable / Registration Concern",
    "Student Fees / Student Financial Services",
    "Learning Services / Math Help / Tutors",
    "Want to Change Program",
    "Connect with Faculty / Program Coordinator / Chair",
    "Health Insurance",
    "Mental Health Support / Counselling",
    "Medical Clinic / Medical Care",
    "Housing / Accommodation",
    "Job Search / Career Services",
    "Immigration / International Student Advising Referral",
    "International Transition Services",
    "International Admissions - Second Program",
    "Library - Tech Loans / TeachMeTech",
    "Library - Research / Writing Consultants",
    "Library - Academic Integrity",
    "CSI - Frosh Kits",
    "CSI - Peer Advocates",
    "Others",
]

# (probability of any count, count function) per category
CAT_PROFILES = {
    "Wayfinding":                                          (0.85, lambda: random.randint(2, 12)),
    "OneCard":                                             (0.55, lambda: random.randint(1, 5)),
    "IT Support":                                          (0.40, lambda: random.randint(1, 3)),
    "Bus Pass / Transportation":                           (0.45, lambda: random.randint(1, 4)),
    "Parking":                                             (0.20, lambda: random.randint(1, 2)),
    "Timetable / Registration Concern":                    (0.40, lambda: random.randint(1, 3)),
    "Student Fees / Student Financial Services":           (0.25, lambda: random.randint(1, 3)),
    "Learning Services / Math Help / Tutors":              (0.20, lambda: random.randint(1, 2)),
    "Want to Change Program":                              (0.10, lambda: 1),
    "Connect with Faculty / Program Coordinator / Chair":  (0.15, lambda: random.randint(1, 2)),
    "Health Insurance":                                    (0.20, lambda: random.randint(1, 2)),
    "Mental Health Support / Counselling":                 (0.18, lambda: random.randint(1, 2)),
    "Medical Clinic / Medical Care":                       (0.10, lambda: 1),
    "Housing / Accommodation":                             (0.15, lambda: random.randint(1, 2)),
    "Job Search / Career Services":                        (0.18, lambda: random.randint(1, 2)),
    "Immigration / International Student Advising Referral": (0.20, lambda: random.randint(1, 3)),
    "International Transition Services":                   (0.12, lambda: random.randint(1, 2)),
    "International Admissions - Second Program":           (0.08, lambda: 1),
    "Library - Tech Loans / TeachMeTech":                  (0.15, lambda: random.randint(1, 3)),
    "Library - Research / Writing Consultants":            (0.12, lambda: random.randint(1, 2)),
    "Library - Academic Integrity":                        (0.08, lambda: 1),
    "CSI - Frosh Kits":                                    (0.10, lambda: random.randint(1, 4)),
    "CSI - Peer Advocates":                                (0.08, lambda: 1),
    "Others":                                              (0.20, lambda: random.randint(1, 2)),
}

OTHERS_INQUIRY_SAMPLES = [
    "Lost & found",
    "Class location for visiting student",
    "Power outlet location",
    "Tour request",
    "Vending machine refund",
    "Lost ID card replacement",
]


VISITOR_HEADERS = [
    "Id",
    "Start time",
    "Completion time",
    "Email",
    "Name",
    "Campus",
    "Time block this submission covers",
    "How many people did you help during this time block?",
    *CATEGORIES,
    "Others (Inquiry)",
]


def weighted_choice(options, weights):
    return random.choices(options, weights=weights, k=1)[0]


def random_datetime_in_block(date, block_label):
    start_h, end_h = TIME_BLOCK_HOURS[block_label]
    h = random.randint(start_h, max(start_h, end_h - 1))
    m = random.randint(0, 59)
    return datetime(date.year, date.month, date.day, h, m)


def fmt_dt(dt):
    return dt.strftime("%Y-%m-%d %H:%M:%S")


def fmt_date(d):
    return d.strftime("%Y-%m-%d")


def generate_visitor_rows(n=40, start=datetime(2026, 3, 16), end=datetime(2026, 5, 4)):
    rows = []
    for _ in range(n):
        days_offset = random.randint(0, (end - start).days)
        submission_date = start + timedelta(days=days_offset)
        # 80% of submissions on weekdays
        if submission_date.weekday() >= 5 and random.random() < 0.8:
            shift = (7 - submission_date.weekday()) % 7
            submission_date += timedelta(days=shift if shift > 0 else 1)

        campus = weighted_choice(CAMPUSES, CAMPUS_WEIGHTS)
        time_block = weighted_choice(TIME_BLOCKS, TIME_BLOCK_WEIGHTS)
        volunteer_name, volunteer_email = random.choice(VOLUNTEERS)

        completion_dt = random_datetime_in_block(submission_date, time_block)
        start_dt = completion_dt - timedelta(minutes=random.randint(2, 5))

        cat_counts = {}
        for cat in CATEGORIES:
            prob, count_fn = CAT_PROFILES[cat]
            cat_counts[cat] = count_fn() if random.random() < prob else 0

        total_inquiries = sum(cat_counts.values())
        if total_inquiries == 0:
            headcount = 0
            others_inquiry = ""
        else:
            headcount = max(1, int(total_inquiries * random.uniform(0.6, 0.9)))
            others_inquiry = (
                random.choice(OTHERS_INQUIRY_SAMPLES) if cat_counts["Others"] > 0 else ""
            )

        row = [
            None,  # placeholder Id, filled after sort
            fmt_dt(start_dt),
            fmt_dt(completion_dt),
            volunteer_email,
            volunteer_name,
            campus,
            time_block,
            headcount,
            *[cat_counts[c] for c in CATEGORIES],
            others_inquiry,
        ]
        rows.append(row)

    # Sort by completion time, then renumber Id
    rows.sort(key=lambda r: r[2])
    for i, r in enumerate(rows):
        r[0] = i + 1
    return rows


# Outreach: hand-crafted, realistic events around real awareness dates
def generate_outreach_rows():
    raw = [
        ("Bell Let's Talk",            "Doon",      datetime(2026, 1, 28, 11, 0), 45,
         "Big atrium turnout. Counselling intake doubled.",
         "Maya Patel", "maya.patel@conestogac.on.ca"),
        ("Bell Let's Talk",            "Waterloo",  datetime(2026, 1, 28, 14, 30), 22,
         "Smaller crowd than Doon but quality conversations.",
         "Jordan Lee", "jordan.lee@conestogac.on.ca"),
        ("Black History Month",        "Doon",      datetime(2026, 2, 14, 12, 15), 52,
         "Speaker drew a packed room.",
         "Aaron Chen", "aaron.chen@conestogac.on.ca"),
        ("Black History Month",        "Reuter",    datetime(2026, 2, 21, 13, 0),  18,
         "",
         "Priya Singh", "priya.singh@conestogac.on.ca"),
        ("International Women's Day",  "Doon",      datetime(2026, 3, 8, 11, 30),  67,
         "Ran out of pins by 1pm. Need 2x next year.",
         "Maya Patel", "maya.patel@conestogac.on.ca"),
        ("International Women's Day",  "Waterloo",  datetime(2026, 3, 8, 13, 45),  24,
         "",
         "Fatima Khan", "fatima.khan@conestogac.on.ca"),
        ("Celebrating Diversity",      "Cambridge", datetime(2026, 3, 21, 14, 0),  31,
         "Food truck collaboration was a hit.",
         "Lucas Tremblay", "lucas.tremblay@conestogac.on.ca"),
        ("Health and Wellness Outreach", "Doon",    datetime(2026, 4, 10, 12, 30), 28,
         "Mental health resources flew off the table.",
         "Sam Rivera", "sam.rivera@conestogac.on.ca"),
        ("Health and Wellness Outreach", "Waterloo", datetime(2026, 4, 12, 13, 15), 19,
         "",
         "Jordan Lee", "jordan.lee@conestogac.on.ca"),
        ("CCR and SSP Promotion",      "Reuter",    datetime(2026, 4, 18, 11, 0),  15,
         "Targeted promo for SSP Spring intake.",
         "Priya Singh", "priya.singh@conestogac.on.ca"),
        ("Sexual Health Week",         "Doon",      datetime(2026, 4, 24, 12, 0),  38,
         "Free supplies running low — restock for fall.",
         "Aaron Chen", "aaron.chen@conestogac.on.ca"),
        ("Sexual Health Week",         "Cambridge", datetime(2026, 4, 26, 14, 30), 12,
         "",
         "Lucas Tremblay", "lucas.tremblay@conestogac.on.ca"),
    ]

    rows = []
    for i, (theme, campus, completion_dt, helped, _notes, name, email) in enumerate(raw, 1):
        start_dt = completion_dt - timedelta(minutes=random.randint(3, 8))
        rows.append([
            i,
            fmt_dt(start_dt),
            fmt_dt(completion_dt),
            email,
            name,
            campus,
            fmt_date(completion_dt.date()),
            helped,
            theme,
        ])
    return rows


def write_tsv(path, headers, rows):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, delimiter="\t")
        writer.writerow(headers)
        writer.writerows(rows)
    print(f"Wrote {path}  ({len(rows)} rows)")


def write_xlsx(path, headers, rows, sheet_name="Sheet1", date_columns=()):
    """Mimic the Forms 'Open in Excel' export: a single sheet with the data
    in a structured table. Power Query reads this exactly like a real
    Forms responses file."""
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Header row
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)

    # Data rows — parse date columns into datetimes so Excel sees them as dates
    date_col_indexes = {headers.index(c) for c in date_columns if c in headers}
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row):
            if col_idx in date_col_indexes and isinstance(value, str):
                # Try datetime first, then date
                try:
                    parsed = datetime.strptime(value, "%Y-%m-%d %H:%M:%S")
                except ValueError:
                    try:
                        parsed = datetime.strptime(value, "%Y-%m-%d").date()
                    except ValueError:
                        parsed = value
                ws.cell(row=row_idx, column=col_idx + 1, value=parsed)
            else:
                ws.cell(row=row_idx, column=col_idx + 1, value=value)

    # Column widths and date formats
    for col_idx, h in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max(12, min(40, len(h) + 2))
        if col_idx - 1 in date_col_indexes:
            for row_idx in range(2, len(rows) + 2):
                ws.cell(row=row_idx, column=col_idx).number_format = (
                    "yyyy-mm-dd hh:mm" if "time" in h.lower() else "yyyy-mm-dd"
                )

    # Wrap in an Excel table — Power Query and pivots like this
    last_col = get_column_letter(len(headers))
    last_row = len(rows) + 1
    table = Table(displayName=sheet_name, ref=f"A1:{last_col}{last_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True
    )
    ws.add_table(table)
    ws.freeze_panes = "A2"

    wb.save(path)
    print(f"Wrote {path}  ({len(rows)} rows)")


def main():
    repo_root = Path(__file__).resolve().parent.parent
    out_dir = repo_root / "templates" / "test-data"

    visitor_rows = generate_visitor_rows()
    outreach_rows = generate_outreach_rows()

    outreach_headers = [
        "Id",
        "Start time",
        "Completion time",
        "Email",
        "Name",
        "Campus",
        "Date of outreach activity",
        "How many people attended the outreach activity?",
        "Outreach Activity",
    ]

    write_tsv(out_dir / "visitor-test-data.tsv", VISITOR_HEADERS, visitor_rows)
    write_tsv(out_dir / "outreach-test-data.tsv", outreach_headers, outreach_rows)

    # XLSX versions for Power Query repointing — sheet name / table name = "Sheet1"
    # so the existing query just sees the same shape it expects.
    write_xlsx(
        out_dir / "Visitor Test Data.xlsx",
        VISITOR_HEADERS,
        visitor_rows,
        sheet_name="Sheet1",
        date_columns=("Start time", "Completion time"),
    )
    write_xlsx(
        out_dir / "Outreach Test Data.xlsx",
        outreach_headers,
        outreach_rows,
        sheet_name="Sheet1",
        date_columns=("Start time", "Completion time", "Date of outreach activity"),
    )

    # Quick sanity
    total_helped = sum(r[7] for r in visitor_rows)
    total_inq = sum(sum(r[8:8 + len(CATEGORIES)]) for r in visitor_rows)
    print(f"\nSummary:")
    print(f"  Visitor: {len(visitor_rows)} submissions, {total_helped} people helped, {total_inq} inquiries logged")
    print(f"  Outreach: {len(outreach_rows)} events, {sum(r[7] for r in outreach_rows)} attended")


if __name__ == "__main__":
    main()
