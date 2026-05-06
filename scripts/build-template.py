#!/usr/bin/env python3
"""
Build the starter Excel dashboard template for Option A+ (v2 — long-format).

Schema matches the user's actual Microsoft Forms export shape (with the
question-text column names like "How many people did you help during this
time block?" intact). Categories are unpivoted into a long-format
`tblVisitorLong` table by Power Query so the dashboard's "Inquiries logged"
KPI and Top Categories chart auto-handle new categories without formula
changes.

Usage (from repo root):
    python scripts/build-template.py

Output:
    templates/student-engagement-dashboard-starter.xlsx
"""

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ---------- Style palette ----------
TEAL = "036C70"
TEAL_DARK = "024C4F"
TEAL_MID = "5CBAB1"
GREY_BG = "F7FAFA"
GREY_BORDER = "D1D1D1"
GREY_TEXT = "707070"
TEXT = "242424"
WHITE = "FFFFFF"

# ---------- Domain constants ----------
CAMPUSES = ["Waterloo", "Doon", "Reuter", "Cambridge"]

# Time-block label values are fuzzy — use a wildcard match in formulas so
# changes to the option text (e.g., en-dash vs hyphen, parens variations)
# don't break the heatmap. The list here drives both the heatmap column
# count and the wildcard prefix used by SUMIFS.
TIME_BLOCK_PREFIXES = ["Morning", "Afternoon"]

# Forms exports the column names exactly as the user typed the questions.
# This list mirrors the user's current form (ASCII hyphens, "Wayfinding"
# without "– General"). The dashboard never sums these as 24 individual
# columns — the long-format table handles that — so adding a new category
# in Forms is a no-op for the dashboard.
CATEGORIES = [
    "Wayfinding",
    "OneCard",
    "IT Support",
    "Bus Pass / Transportation",
    "Parking",
    "Timetable / Registration Concern",
    "Student Fees / Student Financial Services",
    "Learning Services / Math Help / Tutors",
    "Want to Change Program",
    "Connect with Faculty / Program Coordinator / Chair",
    "Health Insurance",
    "Mental Health Support / Counselling",
    "Medical Clinic / Medical Care",
    "Housing / Accommodation",
    "Job Search / Career Services",
    "Immigration / International Student Advising Referral",
    "International Transition Services",
    "International Admissions - Second Program",
    "Library - Tech Loans / TeachMeTech",
    "Library - Research / Writing Consultants",
    "Library - Academic Integrity",
    "CSI - Frosh Kits",
    "CSI - Peer Advocates",
    "Others",
]

# Visitor wide table — matches the Forms export shape exactly.
HEADCOUNT_COL = "How many people did you help during this time block?"
TIME_BLOCK_COL = "Time block this submission covers"

VISITOR_COLS = [
    "Id",
    "Start time",
    "Completion time",
    "Email",
    "Name",
    "Campus",
    TIME_BLOCK_COL,
    HEADCOUNT_COL,
    *CATEGORIES,
    "Others (Inquiry)",
]

# Visitor long-format table — what Power Query produces after Unpivoting
# the 24 category columns. Used by the Top Categories chart and the
# "Inquiries logged" KPI so new categories show up automatically.
VISITOR_LONG_COLS = [
    "Id",
    "Completion time",
    "Email",
    "Name",
    "Campus",
    TIME_BLOCK_COL,
    "Category",
    "Count",
]

# Outreach wide table — matches the Forms export shape exactly.
OUTREACH_HEADCOUNT_COL = "How many people attended the outreach activity?"
OUTREACH_DATE_COL = "Date of outreach activity"

OUTREACH_COLS = [
    "Id",
    "Start time",
    "Completion time",
    "Email",
    "Name",
    "Campus",
    OUTREACH_DATE_COL,
    OUTREACH_HEADCOUNT_COL,
    "Outreach Activity",
]


# ---------- Style helpers ----------
def fill(color):
    return PatternFill("solid", fgColor=color)


def border(color=GREY_BORDER):
    s = Side(border_style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


HEADER_FONT = Font(name="Calibri", size=11, bold=True, color=WHITE)
HEADER_FILL = fill(TEAL)
TITLE_FONT = Font(name="Calibri", size=20, bold=True, color=TEAL_DARK)
SECTION_FONT = Font(name="Calibri", size=12, bold=True, color=WHITE)
KPI_VALUE_FONT = Font(name="Calibri", size=22, bold=True, color=TEAL_DARK)
KPI_LABEL_FONT = Font(name="Calibri", size=9, bold=True, color=GREY_TEXT)
HELP_TEXT_FONT = Font(name="Calibri", size=10, color=GREY_TEXT, italic=True)
BODY_FONT = Font(name="Calibri", size=11, color=TEXT)


def add_section_header(ws, row, text, span=8):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = SECTION_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = 22


def kpi_tile(ws, row, col, label, formula, sub_formula=None):
    title = ws.cell(row=row, column=col, value=label)
    title.font = KPI_LABEL_FONT
    title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    title.fill = fill(GREY_BG)
    title.border = Border(top=Side(border_style="medium", color=TEAL))
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 2)

    val = ws.cell(row=row + 1, column=col, value=formula)
    val.font = KPI_VALUE_FONT
    val.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    val.fill = fill(GREY_BG)
    ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 2)

    sub = ws.cell(row=row + 2, column=col, value=sub_formula or "")
    sub.font = HELP_TEXT_FONT
    sub.alignment = Alignment(horizontal="left", vertical="top", indent=1)
    sub.fill = fill(GREY_BG)
    ws.merge_cells(start_row=row + 2, start_column=col, end_row=row + 2, end_column=col + 2)

    ws.row_dimensions[row].height = 16
    ws.row_dimensions[row + 1].height = 30
    ws.row_dimensions[row + 2].height = 16


# ---------- Sheet builders ----------
def build_data_sheet(ws, columns, table_name, date_columns):
    """Empty wide-format data table with placeholder rows.

    Adds 3 placeholder rows with distinct dates so Excel can build a
    Timeline. Power Query's "Replace" wipes them on first real load.
    """
    sentinel_dates = [
        datetime(2024, 1, 1),
        datetime(2024, 6, 15),
        datetime(2024, 12, 31),
    ]

    for idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = border()

    n_rows = len(sentinel_dates)
    for row_offset, sentinel_date in enumerate(sentinel_dates):
        row = 2 + row_offset
        for idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=row, column=idx)
            if col_name in date_columns:
                cell.value = sentinel_date
                cell.number_format = "yyyy-mm-dd hh:mm" if "time" in col_name.lower() else "yyyy-mm-dd"
            elif col_name == HEADCOUNT_COL or col_name == OUTREACH_HEADCOUNT_COL or col_name in CATEGORIES:
                cell.value = 0
            elif col_name == "Id":
                cell.value = row_offset + 1
            else:
                cell.value = None

    last_col = get_column_letter(len(columns))
    table = Table(displayName=table_name, ref=f"A1:{last_col}{1 + n_rows}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True
    )
    ws.add_table(table)

    for idx, col_name in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = max(14, min(50, len(col_name) + 4))

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    return ws


def build_visitor_long_sheet(ws):
    """Empty long-format table with placeholder rows for Timeline support."""
    sentinel_dates = [
        datetime(2024, 1, 1),
        datetime(2024, 6, 15),
        datetime(2024, 12, 31),
    ]

    columns = VISITOR_LONG_COLS
    for idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = border()

    # 3 placeholder rows, one per sentinel date, with dummy category data
    for row_offset, sentinel_date in enumerate(sentinel_dates):
        row = 2 + row_offset
        ws.cell(row=row, column=1, value=row_offset + 1)  # Id
        ws.cell(row=row, column=2, value=sentinel_date)   # Completion time
        ws.cell(row=row, column=2).number_format = "yyyy-mm-dd hh:mm"
        ws.cell(row=row, column=3, value=None)            # Email
        ws.cell(row=row, column=4, value=None)            # Name
        ws.cell(row=row, column=5, value=None)            # Campus
        ws.cell(row=row, column=6, value=None)            # Time block
        ws.cell(row=row, column=7, value="(placeholder)")  # Category
        ws.cell(row=row, column=8, value=0)                # Count

    last_col = get_column_letter(len(columns))
    table = Table(displayName="tblVisitorLong", ref=f"A1:{last_col}4")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)

    for idx, col_name in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = max(14, min(50, len(col_name) + 4))

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    ws.sheet_state = "hidden"
    return ws


def build_dashboard_sheet(ws):
    ws.sheet_view.showGridLines = False
    for c in range(1, 13):
        ws.column_dimensions[get_column_letter(c)].width = 16

    # Title
    ws["A1"] = "Student Engagement Dashboard"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:H1")
    ws.row_dimensions[1].height = 32

    ws["A2"] = (
        "Lifetime KPIs and heatmap. Slicers/timeline filter the Reports — Weekly / Monthly sheets."
    )
    ws["A2"].font = HELP_TEXT_FONT
    ws.merge_cells("A2:H2")
    ws.row_dimensions[2].height = 18

    # KPI tiles
    add_section_header(ws, 3, "  KEY METRICS")

    kpi_tile(
        ws, 4, 1,
        "PEOPLE HELPED (VISITOR)",
        f"=IFERROR(SUM(tblVisitor[{HEADCOUNT_COL}]),0)",
        "=IFERROR(ROWS(tblVisitor[Id])&\" submissions\",\"No data yet\")",
    )
    kpi_tile(
        ws, 4, 4,
        "INQUIRIES LOGGED",
        "=IFERROR(SUM(tblVisitorLong[Count]),0)",
        "Sum across all categories (auto-includes new ones)",
    )
    kpi_tile(
        ws, 4, 7,
        "TOP CAMPUS BY HEADCOUNT",
        "=IFERROR(INDEX(_Helpers!A4:A7,MATCH(MAX(_Helpers!B4:B7),_Helpers!B4:B7,0)),\"—\")",
        "=IFERROR(MAX(_Helpers!B4:B7)&\" people helped\",\"\")",
    )
    kpi_tile(
        ws, 4, 10,
        "PEOPLE HELPED (OUTREACH)",
        f"=IFERROR(SUM(tblOutreach[{OUTREACH_HEADCOUNT_COL}]),0)",
        "=IFERROR(ROWS(tblOutreach[Id])&\" outreach entries\",\"\")",
    )

    # Heatmap
    add_section_header(ws, 7, "  ENGAGEMENT HEATMAP — CAMPUS × TIME BLOCK")

    ws.cell(row=8, column=1, value="Campus").font = HEADER_FONT
    ws.cell(row=8, column=1).fill = HEADER_FILL
    ws.cell(row=8, column=1).alignment = Alignment(horizontal="center", vertical="center")
    for i, label in enumerate(TIME_BLOCK_PREFIXES):
        c = ws.cell(row=8, column=2 + i, value=label)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")

    for r, campus in enumerate(CAMPUSES):
        rr = 9 + r
        c = ws.cell(row=rr, column=1, value=campus)
        c.font = Font(name="Calibri", size=11, bold=True, color=TEXT)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.fill = fill(GREY_BG)

        for ti, prefix in enumerate(TIME_BLOCK_PREFIXES):
            cell = ws.cell(row=rr, column=2 + ti)
            # Wildcard match — resilient to en-dash vs hyphen, parens variations
            cell.value = (
                f"=IFERROR(SUMIFS(tblVisitor[{HEADCOUNT_COL}],"
                f"tblVisitor[Campus],\"{campus}\","
                f"tblVisitor[{TIME_BLOCK_COL}],\"{prefix}*\"),0)"
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(name="Calibri", size=12, bold=True, color=TEXT)
            cell.border = border()
        ws.row_dimensions[rr].height = 28

    rule = ColorScaleRule(
        start_type="num", start_value=0, start_color=WHITE,
        mid_type="percentile", mid_value=50, mid_color=TEAL_MID,
        end_type="max", end_color=TEAL,
    )
    last_heatmap_col = get_column_letter(1 + len(TIME_BLOCK_PREFIXES))
    ws.conditional_formatting.add(f"B9:{last_heatmap_col}12", rule)

    # Section dividers — actual chart objects get added by post-build.py.
    add_section_header(ws, 14, "  TOP INQUIRY CATEGORIES — auto-populated from tblVisitorLong")
    add_section_header(ws, 30, "  OUTREACH THEMES — PEOPLE HELPED")
    add_section_header(ws, 46, "  OUTREACH REACH BY CAMPUS")

    # Footer help
    add_section_header(ws, 62, "  WHAT'S NEXT")
    note = ws.cell(row=63, column=1, value=(
        "1) Connect Power Query (3 queries: Visitor, VisitorLong, Outreach). "
        "2) Run scripts/post-build.py to add slicers + report pivots + charts. "
        "3) Add a Timeline manually after data flows in. See excel-dashboard-build-guide.md."
    ))
    note.font = HELP_TEXT_FONT
    note.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A63:H65")
    ws.row_dimensions[63].height = 60

    return ws


def build_helpers_sheet(ws):
    ws.sheet_state = "hidden"
    ws.sheet_view.showGridLines = False

    ws["A1"] = "CAMPUS TOTALS"
    ws["A1"].font = Font(bold=True, color=TEAL_DARK)
    ws.merge_cells("A1:B1")
    ws["A3"] = "Campus"
    ws["B3"] = "People helped"
    ws["A3"].font = HEADER_FONT
    ws["B3"].font = HEADER_FONT
    ws["A3"].fill = HEADER_FILL
    ws["B3"].fill = HEADER_FILL

    for i, campus in enumerate(CAMPUSES):
        ws.cell(row=4 + i, column=1, value=campus)
        ws.cell(
            row=4 + i, column=2,
            value=(
                f"=IFERROR(SUMIFS(tblVisitor[{HEADCOUNT_COL}],"
                f"tblVisitor[Campus],\"{campus}\"),0)"
            ),
        )

    for c in (1, 2):
        ws.column_dimensions[get_column_letter(c)].width = 24

    return ws


def build_how_to_use_sheet(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 100

    rows = [
        ("How to use this workbook (v2 — long-format)", "title"),
        ("", "blank"),
        ("This template is built around the actual Microsoft Forms export shape, with category data unpivoted into a long-format table so adding a new question in Forms doesn't break anything.", "body"),
        ("", "blank"),
        ("Sheets", "h2"),
        ("• Dashboard — Shannon's main view. KPI tiles, heatmap, slicers, charts.", "body"),
        ("• Visitor — wide table tblVisitor (one row per submission). Columns match Forms export exactly.", "body"),
        ("• Outreach — wide table tblOutreach. Same.", "body"),
        ("• Reports — Weekly / Monthly — pre-built pivots grouped by date. Filtered by the Dashboard slicers.", "body"),
        ("• VisitorLong (hidden) — long-format table tblVisitorLong with one row per (submission × category). Auto-handles new categories.", "body"),
        ("• _Helpers (hidden) — small helper formulas the Top Campus KPI uses.", "body"),
        ("• _PivotHost (hidden, added by post-build) — owns the slicer-source pivot.", "body"),
        ("", "blank"),
        ("Power Query setup — three queries", "h2"),
        ("On first use, after saving this template into OneDrive:", "body"),
        ("(1) Visitor query — Data → Get Data → From File → From Workbook → pick your Visitor responses .xlsx → Sheet1 → Close & Load To... → Existing worksheet → A1 of Visitor sheet → click Replace when Excel asks → table named tblVisitor.", "body"),
        ("(2) Outreach query — same as Visitor, but pick your Outreach responses .xlsx and load to A1 of Outreach sheet → table named tblOutreach.", "body"),
        ("(3) VisitorLong query — Right-click the Visitor query in the Queries & Connections panel → Reference. New query opens. Rename to VisitorLong. Select the 24 category columns (Wayfinding through Others; NOT Others (Inquiry)). Right-click selected → Unpivot Columns. Rename Attribute → Category and Value → Count. Close & Load To... → Existing worksheet → A1 of VisitorLong sheet (it's hidden — unhide briefly via right-click any tab → Unhide → VisitorLong). Replace tblVisitorLong when prompted. Re-hide.", "body"),
        ("(4) Set all 3 queries to refresh on open + every 5 minutes: Data → Queries & Connections → right-click each query → Properties → tick both checkboxes.", "body"),
        ("", "blank"),
        ("Why this is future-proof", "h2"),
        ("• Adding a new category in Forms (e.g., 'Voter Registration'): Power Query loads it as a new column on tblVisitor. The next refresh of the VisitorLong query auto-includes it via Unpivot Columns. The Top Categories chart and Inquiries Logged KPI both pull from tblVisitorLong, so the new category appears with zero formula edits.", "body"),
        ("• Renaming a question in Forms: Power Query throws on next refresh because it expects the old column name. Edit the query → 'Renamed Columns' step → update the mapping. ~30 seconds.", "body"),
        ("• Adding a new campus: appears automatically in slicers and pivots. Update the helper table on _Helpers (rows 4-7) for the Top Campus KPI.", "body"),
        ("", "blank"),
        ("If something looks broken", "h2"),
        ("• KPI shows #REF! → a column reference is broken. Click the cell to see which column it's looking for, then check that column exists in your data with that exact name.", "body"),
        ("• Heatmap is all zeros even with data → the Time block question wording must START with 'Morning' / 'Afternoon' / 'Evening'. Heatmap formulas use wildcard matching to be resilient, but if your Time block options start with different words, update the heatmap formulas on Dashboard rows 9-12.", "body"),
        ("• Slicers don't filter the Top Categories or Outreach charts → expected. Those charts source from different tables (tblVisitorLong, tblOutreach), so they always show lifetime totals. The slicers/timeline filter only the heatmap-source pivot and the Reports sheets.", "body"),
    ]

    style_map = {
        "title": Font(name="Calibri", size=20, bold=True, color=TEAL_DARK),
        "h2": Font(name="Calibri", size=13, bold=True, color=TEAL),
        "body": Font(name="Calibri", size=11, color=TEXT),
        "blank": None,
    }

    for i, (text, kind) in enumerate(rows):
        cell = ws.cell(row=i + 1, column=2, value=text)
        if style_map[kind]:
            cell.font = style_map[kind]
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if kind == "title":
            ws.row_dimensions[i + 1].height = 32
        elif kind == "h2":
            ws.row_dimensions[i + 1].height = 22
        elif kind == "body":
            ws.row_dimensions[i + 1].height = max(20, 16 + len(text) // 80 * 14)

    return ws


def build_reports_sheet(ws, title, group_hint):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 100

    ws["B1"] = title
    ws["B1"].font = Font(name="Calibri", size=18, bold=True, color=TEAL_DARK)
    ws.row_dimensions[1].height = 30

    ws["B3"] = (
        "PivotTable below is built by post-build.py. "
        f"After data loads, right-click any date → Group → tick {group_hint} for proper rollup."
    )
    ws["B3"].font = HELP_TEXT_FONT
    ws["B3"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("B3:H7")
    ws.row_dimensions[3].height = 70
    return ws


# ---------- Workbook assembly ----------
def main():
    wb = Workbook()
    wb.remove(wb.active)

    how_to_use = wb.create_sheet("How to use")
    dashboard = wb.create_sheet("Dashboard")
    visitor = wb.create_sheet("Visitor")
    outreach = wb.create_sheet("Outreach")
    visitor_long = wb.create_sheet("VisitorLong")
    weekly = wb.create_sheet("Reports — Weekly")
    monthly = wb.create_sheet("Reports — Monthly")
    helpers = wb.create_sheet("_Helpers")

    visitor_dates = ("Start time", "Completion time")
    outreach_dates = ("Start time", "Completion time", OUTREACH_DATE_COL)

    build_data_sheet(visitor, VISITOR_COLS, "tblVisitor", visitor_dates)
    build_data_sheet(outreach, OUTREACH_COLS, "tblOutreach", outreach_dates)
    build_visitor_long_sheet(visitor_long)
    build_helpers_sheet(helpers)
    build_dashboard_sheet(dashboard)
    build_how_to_use_sheet(how_to_use)
    build_reports_sheet(weekly, "Weekly Report", "Days (7) + Months + Years")
    build_reports_sheet(monthly, "Monthly Report", "Months + Years")

    wb.active = wb.sheetnames.index("How to use")

    repo_root = Path(__file__).resolve().parent.parent
    out_path = repo_root / "templates" / "student-engagement-dashboard-starter.xlsx"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Wrote {out_path}")


if __name__ == "__main__":
    main()
